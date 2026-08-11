import sqlite3
from pathlib import Path
from openpyxl import Workbook


# ============================================
# Database
# ============================================

BASE_DIR = Path(__file__).resolve().parent.parent

DB_PATH = BASE_DIR / "database" / "festival.db"


# ============================================
# Database Connection
# ============================================

def get_connection():

    return sqlite3.connect(DB_PATH)


# ============================================
# Dashboard Summary
# ============================================

def get_dashboard_summary():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT COUNT(*) FROM registrations"
    )
    registrations = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM cultural_registrations"
    )
    cultural = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM volunteers"
    )
    volunteers = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM donations"
    )
    donors = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM annaprasada_bookings"
    )
    annaprasada = cursor.fetchone()[0]

    cursor.execute(
        """
        SELECT COALESCE(
            SUM(CAST(amount AS REAL)),
            0
        )
        FROM donations
        """
    )

    donation_amount = cursor.fetchone()[0]

    conn.close()

    return {
        "registrations": registrations,
        "cultural_participants": cultural,
        "volunteers": volunteers,
        "donors": donors,
        "annaprasada_bookings": annaprasada,
        "total_donation_amount": donation_amount
    }


# ============================================
# Generic Table Fetch
# ============================================

def get_table_data(table_name):

    allowed_tables = {

        "registrations":
            "registrations",

        "cultural":
            "cultural_registrations",

        "volunteers":
            "volunteers",

        "donations":
            "donations",

        "annaprasada":
            "annaprasada_bookings"
    }

    if table_name not in allowed_tables:

        raise ValueError(
            "Invalid table name"
        )

    actual_table = allowed_tables[
        table_name
    ]

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        f"""
        SELECT *
        FROM {actual_table}
        ORDER BY id DESC
        """
    )

    rows = cursor.fetchall()

    columns = [
        description[0]
        for description in cursor.description
    ]

    conn.close()

    return columns, rows

# ============================================
# Get Single Admin Record
# ============================================

def get_record_by_id(table_name, record_id):

    allowed_tables = {

        "registrations":
            "registrations",

        "cultural":
            "cultural_registrations",

        "volunteers":
            "volunteers",

        "donations":
            "donations",

        "annaprasada":
            "annaprasada_bookings"
    }


    if table_name not in allowed_tables:

        raise ValueError(
            "Invalid table name"
        )


    actual_table = allowed_tables[
        table_name
    ]


    conn = get_connection()

    cursor = conn.cursor()


    cursor.execute(
        f"""
        SELECT *
        FROM {actual_table}
        WHERE id = ?
        """,
        (record_id,)
    )


    row = cursor.fetchone()


    columns = [
        description[0]
        for description in cursor.description
    ]


    conn.close()


    if row is None:

        return None


    return dict(
        zip(
            columns,
            row
        )
    )

# ============================================
# Update Admin Record
# ============================================

def update_record(
    table_name,
    record_id,
    updates
):

    allowed_tables = {

        "registrations":
            "registrations",

        "cultural":
            "cultural_registrations",

        "volunteers":
            "volunteers",

        "donations":
            "donations",

        "annaprasada":
            "annaprasada_bookings"
    }


    if table_name not in allowed_tables:

        raise ValueError(
            "Invalid table name"
        )


    actual_table = allowed_tables[
        table_name
    ]


    conn = get_connection()

    cursor = conn.cursor()


    # Get actual database columns
    cursor.execute(
        f"PRAGMA table_info({actual_table})"
    )

    table_columns = {
        row[1]
        for row in cursor.fetchall()
    }


    # Never allow ID to be modified
    updates.pop("id", None)


    # Only allow real columns
    safe_updates = {
        key: value
        for key, value in updates.items()
        if key in table_columns
    }


    if not safe_updates:

        conn.close()

        raise ValueError(
            "No valid fields to update"
        )


    set_clause = ", ".join(
        f"{column} = ?"
        for column in safe_updates
    )


    values = list(
        safe_updates.values()
    )

    values.append(record_id)


    cursor.execute(
        f"""
        UPDATE {actual_table}
        SET {set_clause}
        WHERE id = ?
        """,
        values
    )


    if cursor.rowcount == 0:

        conn.close()

        return False


    conn.commit()

    conn.close()


    return True

# ============================================
# Delete Admin Record
# ============================================

def delete_record(
    table_name,
    record_id
):

    allowed_tables = {

        "registrations":
            "registrations",

        "cultural":
            "cultural_registrations",

        "volunteers":
            "volunteers",

        "donations":
            "donations",

        "annaprasada":
            "annaprasada_bookings"
    }

    if table_name not in allowed_tables:

        raise ValueError(
            "Invalid table name"
        )

    actual_table = allowed_tables[
        table_name
    ]

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        f"""
        DELETE FROM {actual_table}
        WHERE id = ?
        """,
        (record_id,)
    )

    deleted = cursor.rowcount

    if deleted == 0:

        conn.close()

        return False

    conn.commit()
    conn.close()

    return True
# ============================================
# Excel Export
# ============================================

def create_excel_file(only_table=None):

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    tables = {

        "Registrations":
            "registrations",

        "Cultural Programs":
            "cultural",

        "Volunteers":
            "volunteers",

        "Donations":
            "donations",

        "Annaprasada":
            "annaprasada"
    }

    # ============================================
    # Select tables
    # ============================================

    if only_table:

        selected_tables = {
            sheet_name: table_name
            for sheet_name, table_name
            in tables.items()
            if table_name == only_table
        }

    else:

        selected_tables = tables


    # ============================================
    # Create Excel sheets
    # ============================================

    for sheet_name, table_name in selected_tables.items():

        columns, rows = get_table_data(
            table_name
        )

        worksheet = workbook.create_sheet(
            title=sheet_name
        )

        # Header
        for column_index, column in enumerate(
            columns,
            start=1
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index
            )

            cell.value = column

            cell.font = cell.font.copy(
                bold=True
            )


        # Data
        for row_index, row in enumerate(
            rows,
            start=2
        ):

            for column_index, value in enumerate(
                row,
                start=1
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index
                ).value = value


    return workbook